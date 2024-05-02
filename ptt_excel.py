import requests
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from ckip_transformers import __version__
from ckip_transformers.nlp import CkipWordSegmenter, CkipPosTagger, CkipNerChunker


# Show version
print(__version__)

# Initialize drivers
print("Initializing drivers ... WS")
ws_driver = CkipWordSegmenter(model="albert-base", device=-1)
print("Initializing drivers ... POS")
pos_driver = CkipPosTagger(model="albert-base", device=-1)
print("Initializing drivers ... NER")
ner_driver = CkipNerChunker(model="albert-base", device=-1)
print("Initializing drivers ... all done")
print()
def clean(sentence_ws, sentence_pos):
  short_with_pos = []
  short_sentence = []
  stop_pos = set(['Nep', 'Nh', 'Nb']) # 這 3 種詞性不保留
  for word_ws, word_pos in zip(sentence_ws, sentence_pos):
    # 只留名詞和動詞
    is_N_or_V = word_pos.startswith("V") or word_pos.startswith("N")
    # 去掉名詞裡的某些詞性
    is_not_stop_pos = word_pos not in stop_pos
    # 只剩一個字的詞也不留
    is_not_one_charactor = not (len(word_ws) == 1)
    # 組成串列
    if is_N_or_V and is_not_stop_pos and is_not_one_charactor:
      short_with_pos.append(f"{word_ws}({word_pos})")
      short_sentence.append(f"{word_ws}")
  return (" ".join(short_sentence), " ".join(short_with_pos))
def main(list1):
    text = [list1]
    ws = ws_driver(text)
    pos = pos_driver(ws)
    ner = ner_driver(text)
    ans=[]
    print()
    print('=====')
    for sentence, sentence_ws, sentence_pos, sentence_ner in zip(text, ws, pos, ner):
        """print("原文：")
        print(sentence)"""
        (short, res) = clean(sentence_ws, sentence_pos)
        """print("斷詞後：")
        print(short)
        print("斷詞後+詞性標注：")"""
        #print(short)
        ans.append(short)
    return(ans)
        #print('=====')

def title():
    title=["標題","回文數量","作者名稱","發文日期","獲取時間"]
    w2xlsx("ptt.xlsx", title)

def w2xlsx(file, data):
    try:
        wb = load_workbook(filename=file)
        ws1 = wb["data"]
        ws1.append(data)
        wb.save(file)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.create_sheet("data", 0)
        sheet.append(data)
        wb.save(file)

def scrape_ptt(url):
    try:
        res = requests.get(url, cookies={'over18': '1'})
        soup = BeautifulSoup(res.text, 'html.parser')
        posts = soup.find_all(class_='r-ent')

        for post in posts:
            try:
                answer = []
                current_date_and_time = datetime.now().strftime("%Y%m%d%H%M%S")
                reply_tag = post.find(class_="nrec")
                user_meta = post.find(class_="author")
                update = post.find(class_="date")
                reply_count = reply_tag.text if reply_tag.text else "0"
                answer.append(post.find(class_='title').find('a').text) #標題
                answer.append(reply_count) #回文數量
                answer.append(user_meta.text) #作者名稱
                answer.append(update.text) #發文日期
                answer.append(current_date_and_time) #獲取時間
                w2xlsx("ptt.xlsx", answer)
            except AttributeError:
                continue

        next_page_link = soup.find(class_="btn wide", text='‹ 上頁')
        next_url = "https://www.ptt.cc" + next_page_link.get('href')
        print(next_url)
        return next_url

    except Exception as e:
        print(e)
        return None
def read_excel(file):
    wb = load_workbook(filename = file)
    sheet = wb['data']

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

if __name__ == "__main__":
    url = "https://www.ptt.cc/bbs/Gossiping/index.html"
    title()
    for i in range(10):
        url = scrape_ptt(url)
        print(url)
    list1=[]
    list2=[]
    excel1=read_excel("ptt.xlsx")
    for i in range(1,len(read_excel("ptt.xlsx"))):
        abc=excel1[i][0]
        abc=abc.replace("[問卦] ","")
        abc=abc.replace("Re: ","")
        abc=abc.replace("[新聞] ","")
        abc=abc.replace(" ","")
        list1.append(abc)

    f = open('text.txt', 'w', encoding='utf-8')  # 指定檔案編碼為UTF-8
    for i in range(0,len(list1)):
        list2.append(main(list1[i]))
    for i in range(0,len(list2)):
        for x in range(0,len(list2[i])):
            f.write(list2[i][x])
    f.close()
