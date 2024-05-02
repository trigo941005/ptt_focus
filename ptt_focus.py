from wordcloud import WordCloud, STOPWORDS
import matplotlib.pyplot as plt
from collections import Counter # 次數統計
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from ckip_transformers import __version__
from ckip_transformers.nlp import CkipWordSegmenter, CkipPosTagger, CkipNerChunker
import time
import copy
import os
fontpath = "C:\Windows\Fonts\kaiu.ttf"  # 字型檔


def in_out(going_page,post,before_answer):
    #print(going_page)
    going = requests.get(going_page,cookies={'over18': '1'})
    soup = BeautifulSoup(going.text, 'html.parser')
    #print(soup)
    posts = soup.find_all(id="main-container")
    for push in posts:
        push = push.find_all(class_ = 'push')
        for post in push:
            answer = list()
            answer = copy.deepcopy(before_answer)
            reply_tag = post.find(class_="f1 hl push-tag").text #回文聲量
            reply_userid = post.find(class_="f3 hl push-userid").text #回文作者
            reply_content = post.find(class_='f3 push-content').text#回文內容    #增加.text已取得位置內容
            reply_ipdatetime = post.find(class_='push-ipdatetime').text#回文時間
            answer.append(reply_tag)
            answer.append(reply_userid)
            answer.append(reply_content)
            answer.append(reply_ipdatetime)
            print(answer)
            """back_page_link = soup.find(class_="board").get('href')#.find('a')
            GoOut_page = ("https://www.ptt.cc" +(back_page_link))"""
            w2xlsx("focus.xlsx", answer)
    #return GoOut_page
    #print(reply_userid)

def title():
    title=["標題","回文聲量","回文作者","回文內容","回文時間"]
    w2xlsx("focus.xlsx", title)
def w2xlsx(file, data):
    try:
        wb = load_workbook(filename=file)
        ws1 = wb["data"]
        ws1.append(data)
        wb.save(file)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.create_sheet("data", 0)
        sheet.append(data)
        wb.save(file)
def scrape_ptt(url,focus):
    #try:
    res = requests.get(url, cookies={'over18': '1'})
    soup = BeautifulSoup(res.text, 'html.parser')
    posts = soup.find_all(class_='r-ent')
    #print(posts)
    for post in posts:

        #try:
        answer = []
        try:
            if focus in (post.find(class_='title').find('a').text):
                reply_tag = post.find(class_="nrec")
                reply_count = reply_tag.text if reply_tag.text else "0"
                answer.append(post.find(class_='title').find('a').text) #標題
                #answer.append(reply_count) #回文數量
                going_page=("https://www.ptt.cc" +(post.find(class_='title').find('a')).get('href'))
                in_out(going_page,post,answer)
        except AttributeError as e:
            #print(e)
            continue

    next_page_link = soup.find(class_="btn wide", text='‹ 上頁')
    next_url = "https://www.ptt.cc" + next_page_link.get('href')
    #print(next_url)
    return next_url
# Read the whole text.
path = 'C:\\Users\\88690\\Desktop\\peter\\20240422\\focus.xlsx'
os.remove(path) 
txtfile = "text.txt" # 剛才下載存的文字檔
text = open(txtfile,"r",encoding='utf-8').read()
text=text.split(" ")
dictionary = Counter(text)
freq = {}
for ele in dictionary:
    if ele in text:
        freq[ele] = dictionary[ele]
# Sort dictionary values while keeping keys
sorted_items = sorted(freq.items(), key=lambda x: x[1])

focus=(list(sorted_items[-1])[0])  # Output: [('b', 1), ('c', 2), ('a', 3)]
if __name__ == "__main__":
    url = "https://www.ptt.cc/bbs/Gossiping/index.html"
    title()
    for i in range(10):
        url = scrape_ptt(url,focus)
        #print(url)
wordcloud = WordCloud(background_color="white", contour_width=3, contour_color='steelblue', font_path= fontpath).generate_from_frequencies(freq)